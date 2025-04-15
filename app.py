import pandas as pd
import streamlit as st
from io import BytesIO
import numpy as np
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from joblib import Parallel, delayed
import time

# Ustawienie limitu komórek dla Pandas Styler
pd.set_option("styler.render.max_elements", 3000000)

# Lista kolumn binarnych do sprawdzenia
columns_to_check = [
    'Katalogowa PLN', 'Promocyjna PLN', 'Katalogowa CZK', 'Promocyjna CZK',
    'Katalogowa RON', 'Promocyjna RON', 'Katalogowa BGN', 'Promocyjna BGN',
    'Katalogowa EUR DE', 'Promocyjna EUR DE', 'Katalogowa EUR GR', 'Promocyjna EUR GR',
    'Katalogowa EUR IT', 'Promocyjna EUR IT', 'Katalogowa EUR LT', 'Promocyjna EUR LT',
    'Katalogowa EUR SK', 'Promocyjna EUR SK', 'Katalogowa UAH', 'Promocyjna UAH',
    'Katalogowa HUF', 'Promocyjna HUF', 'Marketplace PL', 'Marketplace CZ',
    'Marketplace RO', 'Kaufland EUR DE', 'Empik PLN', 'Marketplace Amazon DE',
    'Marketplace FR', 'Marketplace IT', 'Marketplace Amazon ES', 'Marketplace HU',
    'Marketplace SK', 'Marketplace SE', 'Marketplace UAH'
]

# Funkcja do wczytywania pliku w partiach
@st.cache_data
def load_file(uploaded_file, usecols, chunksize=10000):
    if uploaded_file is not None:
        try:
            chunks = []
            if uploaded_file.name.endswith('.csv'):
                for chunk in pd.read_csv(uploaded_file, low_memory=False, dtype={'index': str, 'Indeks': str, 'modelcolor': str}, usecols=usecols, chunksize=chunksize):
                    chunks.append(chunk)
            elif uploaded_file.name.endswith('.xlsx'):
                return pd.read_excel(uploaded_file, engine='openpyxl', dtype={'index': str, 'Indeks': str, 'modelcolor': str}, usecols=usecols)
            return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()
        except Exception as e:
            st.error(f"Błąd wczytywania pliku {uploaded_file.name}: {str(e)}")
            return None
    return None

# Funkcja do czyszczenia nazw kolumn
def clean_column_names(df):
    df.columns = df.columns.str.strip()
    return df

# Funkcja do sprawdzania spójności dla jednej kolumny (do równoległego przetwarzania)
def check_column_consistency(df, col):
    # Grupowanie po modelcolor i sprawdzenie spójności
    grouped = df.groupby('modelcolor')[col]
    non_null_counts = grouped.apply(lambda x: x[x.isin([0, 1])].nunique())
    inconsistent_modelcolors = non_null_counts[non_null_counts > 1].index

    if len(inconsistent_modelcolors) == 0:
        return pd.DataFrame()

    # Tworzenie raportu dla niespójnych modelcolor
    inconsistent_df = df[df['modelcolor'].isin(inconsistent_modelcolors)][['modelcolor', 'index', 'Producent', 'Kat 1', 'last_delivery_date']].copy()
    inconsistent_df['problem_column'] = col
    inconsistent_df['problem_value'] = df.loc[inconsistent_df.index, col]
    inconsistent_df['issue'] = np.where(
        inconsistent_df['problem_value'].isin([0, 1]) | inconsistent_df['problem_value'].isna(),
        f"Niespójność w {col} (różne wartości 0/1)",
        f"Niepoprawna wartość w {col} (oczekiwano 0 lub 1)"
    )
    return inconsistent_df

# Zoptymalizowana funkcja do sprawdzania spójności z równoległym przetwarzaniem
@st.cache_data
def check_consistency(df, columns_to_check):
    start_time = time.time()
    st.write(f"Rozpoczęcie sprawdzania spójności: {time.strftime('%H:%M:%S')}")

    # Równoległe sprawdzanie spójności dla każdej kolumny
    results = Parallel(n_jobs=-1)(
        delayed(check_column_consistency)(df, col) for col in columns_to_check
    )

    # Łączenie wyników
    result_df = pd.concat(results, ignore_index=True) if results else pd.DataFrame()
    if not result_df.empty:
        result_df = result_df.sort_values(by=['modelcolor', 'last_delivery_date'])

    end_time = time.time()
    st.write(f"Zakończono sprawdzanie spójności w {end_time - start_time:.2f} sekund: {time.strftime('%H:%M:%S')}")
    return result_df

# Funkcja do podświetlania tylko problematycznej wartości w Streamlit
def highlight_issues(row):
    styles = [''] * len(row)
    if row['issue']:
        for i, col in enumerate(row.index):
            if col == 'problem_value':
                styles[i] = 'background-color: red'
                break
    return styles

# Funkcja do zapisu pliku Excel z podświetleniem
def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Raport')
        if not df.empty:
            workbook = writer.book
            worksheet = writer.sheets['Raport']
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

            for row_idx, row_data in df.iterrows():
                if row_data['issue']:
                    for col_idx, col_name in enumerate(df.columns, start=1):
                        if col_name == 'problem_value':
                            worksheet.cell(row=row_idx + 2, column=col_idx).fill = red_fill
                            break

            for col_idx in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 25

    return output.getvalue()

# Tytuł aplikacji
st.title("Sprawdzanie spójności binarnych danych stałych cen")

# Opcje wczytywania danych
st.subheader("Ustawienia wczytywania danych")
chunk_size = st.number_input("Wczytaj dane w partiach (rozmiar partii, 0 = wczytaj całość):", min_value=0, value=10000, step=1000)
max_rows = st.number_input("Maksymalna liczba wierszy do analizy (0 = bez limitu):", min_value=0, value=50000, step=10000)

# Wczytanie plików
uploaded_file1 = st.file_uploader("Wybierz plik z bazą danych", type=["csv", "xlsx"])
uploaded_file2 = st.file_uploader("Wybierz plik ze stałymi cenami", type=["csv", "xlsx"])

if uploaded_file1 is not None and uploaded_file2 is not None:
    # Definicja potrzebnych kolumn do wczytania
    base_cols = ['index', 'modelcolor', 'last_delivery_date']
    price_cols = ['Indeks', 'Producent', 'Kat 1'] + columns_to_check

    # Wczytanie danych w partiach
    start_time = time.time()
    with st.spinner("Wczytywanie pliku z bazą danych..."):
        df_base = load_file(uploaded_file1, usecols=base_cols, chunksize=chunk_size if chunk_size > 0 else None)
    with st.spinner("Wczytywanie pliku ze stałymi cenami..."):
        df_prices = load_file(uploaded_file2, usecols=price_cols, chunksize=chunk_size if chunk_size > 0 else None)
    st.write(f"Wczytywanie danych trwało {time.time() - start_time:.2f} sekund")

    if df_base is None or df_prices is None:
        st.error("Nie udało się wczytać jednego z plików. Sprawdź format lub zawartość.")
    elif df_base.empty or df_prices.empty:
        st.error("Wczytane pliki są puste. Sprawdź ich zawartość.")
    else:
        # Ograniczenie liczby wierszy, jeśli ustawiono limit
        if max_rows > 0:
            df_base = df_base.head(max_rows)
            df_prices = df_prices.head(max_rows)

        df_base = clean_column_names(df_base)
        df_prices = clean_column_names(df_prices)

        # Sprawdzanie wymaganych kolumn
        missing_base_cols = [col for col in base_cols if col not in df_base.columns]
        missing_price_cols = [col for col in price_cols if col not in df_prices.columns]

        if missing_base_cols or missing_price_cols:
            st.error(f"Brakujące kolumny w pliku z bazą: {missing_base_cols}, w pliku z cenami: {missing_price_cols}")
        else:
            # Normalizacja nazw kolumn
            df_prices = df_prices.rename(columns={'Indeks': 'index'})

            # Łączenie danych
            @st.cache_data
            def merge_data(df_base, df_prices):
                start_time = time.time()
                with st.spinner("Łączenie danych..."):
                    merged_df = pd.merge(df_base[['index', 'modelcolor', 'last_delivery_date']],
                                        df_prices[['index', 'Producent', 'Kat 1'] + columns_to_check],
                                        how='left',
                                        on='index')
                st.write(f"Łączenie danych trwało {time.time() - start_time:.2f} sekund")
                return merged_df

            merged_df = merge_data(df_base, df_prices)

            # Filtrowanie po modelcolor z wyszukiwaniem
            unique_modelcolors = merged_df['modelcolor'].unique()
            st.subheader("Filtrowanie po modelcolor")
            modelcolor_search = st.text_input("Wyszukaj modelcolor (wpisz fragment, aby zawęzić listę):", "")
            if modelcolor_search:
                filtered_modelcolors = [mc for mc in unique_modelcolors if modelcolor_search.lower() in str(mc).lower()]
            else:
                filtered_modelcolors = unique_modelcolors
            selected_modelcolors = st.multiselect("Wybierz modelcolor do analizy (zostaw puste, aby analizować wszystkie)", 
                                                options=filtered_modelcolors, 
                                                default=[])

            # Filtrowanie po Producent z wyszukiwaniem
            unique_producents = merged_df['Producent'].unique()
            st.subheader("Filtrowanie po Producent")
            producent_search = st.text_input("Wyszukaj Producent (wpisz fragment, aby zawęzić listę):", "")
            if producent_search:
                filtered_producents = [p for p in unique_producents if producent_search.lower() in str(p).lower()]
            else:
                filtered_producents = unique_producents
            selected_producents = st.multiselect("Wybierz Producent do analizy (zostaw puste, aby analizować wszystkich)", 
                                                options=filtered_producents, 
                                                default=[])

            # Filtrowanie danych
            filtered_df = merged_df
            if selected_modelcolors:
                filtered_df = filtered_df[filtered_df['modelcolor'].isin(selected_modelcolors)]
            if selected_producents:
                filtered_df = filtered_df[filtered_df['Producent'].isin(selected_producents)]

            # Diagnostyka: Wyświetlenie danych po złączeniu dla wybranego modelcolor
            if selected_modelcolors:
                st.subheader("Dane po złączeniu dla wybranego modelcolor (diagnostyka)")
                for modelcolor in selected_modelcolors:
                    st.write(f"**modelcolor = {modelcolor}**")
                    modelcolor_data = filtered_df[filtered_df['modelcolor'] == modelcolor][['index', 'Producent', 'Kat 1', 'last_delivery_date'] + columns_to_check]
                    st.dataframe(modelcolor_data)

            # Sprawdzanie spójności
            with st.spinner("Sprawdzanie spójności binarnych danych stałych cen..."):
                issues_df = check_consistency(filtered_df, columns_to_check)

            # Wybór liczby wierszy do wyświetlenia
            st.subheader("Ustawienia wyświetlania")
            max_rows = len(issues_df) if not issues_df.empty else 0
            display_rows = st.slider("Wybierz liczbę wierszy do wyświetlenia",
                                    min_value=1, max_value=max_rows, value=min(100, max_rows), step=10) if max_rows > 0 else 1

            if issues_df.empty:
                st.success("Wszystkie binarne dane stałych cen są spójne w obrębie modelcolor!")
            else:
                st.warning(f"Znaleziono {len(issues_df)} niespójności lub niepoprawnych wartości w binarnych danych stałych cen:")
                
                # Wyświetlanie raportu z podświetleniem
                styled_issues = issues_df.head(display_rows).style.apply(highlight_issues, axis=1)
                st.dataframe(styled_issues)

                # Pobieranie pliku z raportem
                if st.button("Pobierz raport z problemami (Excel)"):
                    with st.spinner("Generowanie pliku Excel..."):
                        excel_data = to_excel(issues_df)
                    st.download_button(
                        label="Pobierz raport (Excel)",
                        data=excel_data,
                        file_name="issues_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_excel"
                    )
else:
    st.info("Proszę załadować oba pliki, aby kontynuować.")